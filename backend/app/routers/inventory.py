from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
from pydantic import BaseModel
import io
from datetime import date, timedelta
from app.database import get_db
from app.models.inventory import Ingredient, Recipe, StockMovement
from app.models.menu import MenuItem
from app.schemas.inventory import (IngredientCreate, IngredientUpdate, IngredientOut,
                                    StockMovementCreate, StockMovementOut,
                                    RecipeCreate, RecipeLineIn, RecipeOut,
                                    ShoppingListEntry, ShoppingListLine)
from app.auth.dependencies import get_current_user, admin_only
from app.models.user import User

router = APIRouter(prefix="/api/inventory", tags=["inventory"])


@router.get("/ingredients", response_model=List[IngredientOut])
def list_ingredients(db: Session = Depends(get_db), _=Depends(get_current_user)):
    ingredients = db.query(Ingredient).all()
    result = []
    for ing in ingredients:
        d = {c.name: getattr(ing, c.name) for c in ing.__table__.columns}
        d["bajo_stock"] = ing.stock_actual <= ing.stock_minimo
        result.append(d)
    return result


@router.post("/ingredients", response_model=IngredientOut)
def create_ingredient(data: IngredientCreate, db: Session = Depends(get_db), _=Depends(admin_only)):
    existing = db.query(Ingredient).filter(
        Ingredient.nombre.ilike(data.nombre.strip())
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail=f"Ya existe un insumo con el nombre '{existing.nombre}'")
    ing = Ingredient(**data.model_dump())
    db.add(ing)
    db.commit()
    db.refresh(ing)
    d = {c.name: getattr(ing, c.name) for c in ing.__table__.columns}
    d["bajo_stock"] = ing.stock_actual <= ing.stock_minimo
    return d


@router.delete("/ingredients/{ing_id}")
def delete_ingredient(ing_id: int, db: Session = Depends(get_db), _=Depends(admin_only)):
    ing = db.query(Ingredient).filter(Ingredient.id == ing_id).first()
    if not ing:
        raise HTTPException(status_code=404, detail="Ingrediente no encontrado")
    for r in db.query(Recipe).filter(Recipe.ingredient_id == ing_id).all():
        db.delete(r)
    for m in db.query(StockMovement).filter(StockMovement.ingredient_id == ing_id).all():
        db.delete(m)
    db.flush()
    db.delete(ing)
    db.commit()
    return {"ok": True}


@router.get("/ingredients/{ing_id}/usages")
def get_ingredient_usages(ing_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    """Retorna los platos (recetas) que usan este ingrediente."""
    recipes = db.query(Recipe).filter(Recipe.ingredient_id == ing_id).all()
    return [
        {"recipe_id": r.id, "item_id": r.item_id, "item_nombre": r.item.nombre if r.item else "—", "cantidad": r.cantidad}
        for r in recipes
    ]


@router.delete("/recipes/{recipe_id}")
def delete_recipe_line(recipe_id: int, db: Session = Depends(get_db), _=Depends(admin_only)):
    r = db.query(Recipe).filter(Recipe.id == recipe_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Línea de receta no encontrada")
    db.delete(r)
    db.commit()
    return {"ok": True}


@router.put("/ingredients/{ing_id}", response_model=IngredientOut)
def update_ingredient(ing_id: int, data: IngredientUpdate, db: Session = Depends(get_db), _=Depends(admin_only)):
    ing = db.query(Ingredient).filter(Ingredient.id == ing_id).first()
    if not ing:
        raise HTTPException(status_code=404, detail="Ingrediente no encontrado")
    for field, value in data.model_dump(exclude_none=True).items():
        setattr(ing, field, value)
    db.commit()
    db.refresh(ing)
    d = {c.name: getattr(ing, c.name) for c in ing.__table__.columns}
    d["bajo_stock"] = ing.stock_actual <= ing.stock_minimo
    return d


@router.post("/movements", response_model=StockMovementOut)
def register_movement(data: StockMovementCreate, db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user)):
    ing = db.query(Ingredient).filter(Ingredient.id == data.ingredient_id).first()
    if not ing:
        raise HTTPException(status_code=404, detail="Ingrediente no encontrado")
    if data.tipo == "entrada":
        ing.stock_actual += data.cantidad
    elif data.tipo == "salida":
        if ing.stock_actual < data.cantidad:
            raise HTTPException(status_code=400, detail="Stock insuficiente")
        ing.stock_actual -= data.cantidad
    mov = StockMovement(ingredient_id=data.ingredient_id, tipo=data.tipo,
                        cantidad=data.cantidad, motivo=data.motivo, usuario_id=current_user.id)
    db.add(mov)
    db.commit()
    db.refresh(mov)
    return mov


@router.get("/movements", response_model=List[StockMovementOut])
def list_movements(ingredient_id: int = None, db: Session = Depends(get_db), _=Depends(get_current_user)):
    q = db.query(StockMovement)
    if ingredient_id:
        q = q.filter(StockMovement.ingredient_id == ingredient_id)
    return q.order_by(StockMovement.created_at.desc()).limit(100).all()


@router.get("/recipes/export-excel")
def export_recipes_excel(db: Session = Depends(get_db), _=Depends(get_current_user)):
    """Descarga todas las recetas en un archivo Excel listo para editar."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from app.models.menu import MenuItem

    wb = Workbook()
    ws = wb.active
    ws.title = "Recetas"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="EA580C")
    border_thin = Border(
        left=Side(style="thin", color="E4E4E7"), right=Side(style="thin", color="E4E4E7"),
        top=Side(style="thin", color="E4E4E7"),  bottom=Side(style="thin", color="E4E4E7"),
    )
    center = Alignment(horizontal="center", vertical="center")

    headers = ["Plato", "Ingrediente", "Cantidad", "Unidad"]
    col_widths = [35, 30, 12, 12]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border_thin
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
    ws.row_dimensions[1].height = 22

    items = db.query(MenuItem).order_by(MenuItem.nombre).all()
    row = 2
    for item in items:
        recipes = db.query(Recipe).filter(Recipe.item_id == item.id).all()
        if not recipes:
            continue
        for r in recipes:
            ing = r.ingrediente
            if not ing:
                continue
            unidad = r.unidad if r.unidad else ing.unidad
            ws.cell(row=row, column=1, value=item.nombre).border = border_thin
            ws.cell(row=row, column=2, value=ing.nombre).border = border_thin
            ws.cell(row=row, column=3, value=r.cantidad).border = border_thin
            ws.cell(row=row, column=4, value=unidad).border = border_thin
            ws.cell(row=row, column=3).alignment = center
            ws.cell(row=row, column=4).alignment = center
            row += 1

    wi = wb.create_sheet("Instrucciones")
    instrucciones = [
        ("INSTRUCCIONES PARA EDITAR RECETAS", True),
        ("", False),
        ("1. En la hoja 'Recetas' puedes modificar la Cantidad y la Unidad de cada ingrediente.", False),
        ("2. NO cambies los nombres de las columnas 'Plato' ni 'Ingrediente' (se usan para identificar la receta).", False),
        ("3. Para ELIMINAR un ingrediente de una receta: borra toda la fila.", False),
        ("4. Para AGREGAR un ingrediente nuevo: agrega una fila nueva con el nombre exacto del plato e ingrediente.", False),
        ("5. Las unidades válidas son: g, kg, ml, L, und, taza, cda, cdta", False),
        ("6. Guarda el archivo y súbelo usando el botón 'Importar Excel' en la aplicación.", False),
        ("", False),
        ("IMPORTANTE: Los nombres de Plato e Ingrediente deben coincidir exactamente con los registrados en el sistema.", False),
    ]
    wi.column_dimensions["A"].width = 90
    for i, (text, bold) in enumerate(instrucciones, 1):
        cell = wi.cell(row=i, column=1, value=text)
        cell.font = Font(bold=True, size=13, color="EA580C") if bold else Font(size=11)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=recetas.xlsx"},
    )


def _normalize(s: str) -> str:
    """Minúsculas, sin tildes, sin espacios extra."""
    import unicodedata
    s = s.strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s


def _parse_recipes_excel(file_bytes: bytes, db, clear_all: bool):
    from openpyxl import load_workbook
    from app.models.menu import MenuItem

    try:
        wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
        ws = wb["Recetas"]
    except Exception:
        raise HTTPException(400, "No se pudo leer el archivo. Asegúrate de subir el Excel generado por el sistema.")

    items_by_name = {_normalize(i.nombre): i for i in db.query(MenuItem).all()}
    ings_by_name  = {_normalize(i.nombre): i for i in db.query(Ingredient).all()}

    rows_by_plato: dict[int, list[dict]] = {}
    errors: list[str] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        plato_nombre = str(row[0]).strip()
        ing_nombre   = str(row[1]).strip() if row[1] else ""
        cantidad_raw = row[2]
        unidad       = str(row[3]).strip() if row[3] else ""

        item = items_by_name.get(_normalize(plato_nombre))
        if not item:
            errors.append(f"Plato no encontrado: '{plato_nombre}'")
            continue
        ing = ings_by_name.get(_normalize(ing_nombre))
        if not ing:
            errors.append(f"Ingrediente no encontrado: '{ing_nombre}' (plato: {plato_nombre})")
            continue
        try:
            cantidad = float(cantidad_raw)
        except (TypeError, ValueError):
            errors.append(f"Cantidad inválida: {plato_nombre} / {ing_nombre}")
            continue

        rows_by_plato.setdefault(item.id, []).append({
            "ingredient_id": ing.id,
            "cantidad": cantidad,
            "unidad": unidad or ing.unidad,
        })

    if errors and not rows_by_plato:
        raise HTTPException(422, detail={"mensaje": "No se importó nada.", "errores": errors})

    # Borrar recetas según el modo
    if clear_all:
        for r in db.query(Recipe).all():
            db.delete(r)
        db.flush()

    updated_platos = 0
    for item_id, lines in rows_by_plato.items():
        if not clear_all:
            for r in db.query(Recipe).filter(Recipe.item_id == item_id).all():
                db.delete(r)
            db.flush()
        for line in lines:
            db.add(Recipe(
                item_id=item_id,
                ingredient_id=line["ingredient_id"],
                cantidad=line["cantidad"],
                unidad=line["unidad"],
            ))
        updated_platos += 1

    db.commit()
    return {
        "ok": True,
        "platos_actualizados": updated_platos,
        "lineas_importadas": sum(len(v) for v in rows_by_plato.values()),
        "advertencias": errors,
    }


@router.post("/recipes/import-excel")
def import_recipes_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(admin_only),
):
    """Importa recetas — reemplaza solo los platos presentes en el archivo."""
    return _parse_recipes_excel(file.file.read(), db, clear_all=False)


@router.post("/recipes/replace-all-excel")
def replace_all_recipes_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(admin_only),
):
    """Borra TODAS las recetas y las reemplaza con el contenido del archivo."""
    return _parse_recipes_excel(file.file.read(), db, clear_all=True)


@router.get("/recipes/{item_id}", response_model=List[RecipeOut])
def get_recipes(item_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    recipes = db.query(Recipe).filter(Recipe.item_id == item_id).all()
    result = []
    for r in recipes:
        d = {c.name: getattr(r, c.name) for c in r.__table__.columns}
        d["ingrediente_nombre"] = r.ingrediente.nombre if r.ingrediente else None
        d["unidad"] = r.unidad if r.unidad else (r.ingrediente.unidad if r.ingrediente else None)
        result.append(d)
    return result


@router.post("/recipes", response_model=RecipeOut)
def create_recipe(data: RecipeCreate, db: Session = Depends(get_db), _=Depends(admin_only)):
    recipe = Recipe(**data.model_dump())
    db.add(recipe)
    db.commit()
    db.refresh(recipe)
    d = {c.name: getattr(recipe, c.name) for c in recipe.__table__.columns}
    d["ingrediente_nombre"] = recipe.ingrediente.nombre if recipe.ingrediente else None
    d["unidad"] = recipe.ingrediente.unidad if recipe.ingrediente else None
    return d


@router.put("/recipes/{item_id}/bulk", response_model=List[RecipeOut])
def save_recipes_bulk(item_id: int, lines: List[RecipeLineIn],
                      db: Session = Depends(get_db), _=Depends(admin_only)):
    """Reemplaza TODOS los ingredientes de un plato de una vez."""
    db.query(Recipe).filter(Recipe.item_id == item_id).delete()
    result = []
    for line in lines:
        ing = db.query(Ingredient).filter(Ingredient.id == line.ingredient_id).first()
        if not ing:
            raise HTTPException(400, f"Ingrediente {line.ingredient_id} no existe")
        unidad = line.unidad if line.unidad else ing.unidad
        r = Recipe(item_id=item_id, ingredient_id=line.ingredient_id, cantidad=line.cantidad, unidad=unidad)
        db.add(r)
        db.flush()
        d = {c.name: getattr(r, c.name) for c in r.__table__.columns}
        d["ingrediente_nombre"] = ing.nombre
        d["unidad"] = unidad
        result.append(d)
    db.commit()
    return result


@router.post("/shopping-list", response_model=List[ShoppingListLine])
def shopping_list(entries: List[ShoppingListEntry],
                  db: Session = Depends(get_db), _=Depends(get_current_user)):
    """
    Recibe lista de {item_id, porciones} y devuelve lo que hay que comprar
    comparando recetas vs stock actual.
    """
    needed: dict[int, float] = {}   # ingredient_id → gramos totales

    for entry in entries:
        recipes = db.query(Recipe).filter(Recipe.item_id == entry.item_id).all()
        for r in recipes:
            needed[r.ingredient_id] = needed.get(r.ingredient_id, 0) + r.cantidad * entry.porciones

    result = []
    for ing_id, total in needed.items():
        ing = db.query(Ingredient).filter(Ingredient.id == ing_id).first()
        if not ing:
            continue
        faltante = max(0.0, total - ing.stock_actual)
        result.append(ShoppingListLine(
            ingredient_id=ing_id,
            nombre=ing.nombre,
            unidad=ing.unidad,
            necesario=round(total, 2),
            stock_actual=round(ing.stock_actual, 2),
            faltante=round(faltante, 2),
            costo_estimado=round(faltante * (ing.costo_unitario or 0), 2),
            categoria=ing.categoria or 'Otros',
        ))

    return sorted(result, key=lambda x: -x.faltante)


# ── Info del día siguiente según ciclo activo ─────────────────────────────────

import unicodedata

def _normalizar(s: str) -> str:
    """Convierte a minúsculas y elimina tildes para comparación robusta."""
    return unicodedata.normalize("NFD", s.lower()).encode("ascii", "ignore").decode("ascii").strip()

# Platos que se sirven TODOS los días independientemente del ciclo.
# Formato: (nombre, subtipo)  subtipo = "acompanante" | "proteina_fija"
# Fuente: imágenes reales de los 4 ciclos (33 fotos del tablero diario).
PLATOS_FIJOS: list[tuple[str, str]] = [
    # Acompañantes base (siempre incluidos en el Menú del Día $18.000)
    ("ARROZ BLANCO",        "acompanante"),
    ("ARROZ COCO",          "acompanante"),
    ("FRIJOLES",            "acompanante"),
    # Proteínas generales — exactamente las que aparecen en TODOS los menús:
    # "Pollo Horno / Milanesa de Pollo o de Cerdo / Res - Cerdo - Pollo / Chicharron - Molida"
    ("POLLO AL HORNO",      "proteina_fija"),
    ("MILANESA DE POLLO",   "proteina_fija"),
    ("MILANESA DE CERDO",   "proteina_fija"),
    ("RES A LA PLANCHA",    "proteina_fija"),
    ("CERDO A LA PLANCHA",  "proteina_fija"),
    ("POLLO A LA PLANCHA",  "proteina_fija"),   # "Pollo" en "Res-Cerdo-Pollo"
    ("CHICHARRON",          "proteina_fija"),
    ("CARNE MOLIDA",        "proteina_fija"),
    # Nota: Filete de Pescado Apanado/Plancha es Menú de la Casa ($21.000), NO del día
]

# Días en que se ofrece Consomé de Pescado como SEGUNDA sopa (además de la del ciclo)
# Fuente: confirmado en todas las imágenes de Viernes, Sábado y Domingo de los 4 ciclos
DIAS_CON_CONSOMÉ = {5, 6, 7}  # 5=Viernes, 6=Sábado, 7=Domingo


@router.get("/next-day-info")
def next_day_info(db: Session = Depends(get_db), _=Depends(get_current_user)):
    """
    Devuelve TODOS los platos del ciclo activo para mañana:
      - sopa y proteína del ciclo
      - contornos del día (parseados individualmente desde texto libre)
      - ensalada del día (si tiene ensalada_id)
      - platos fijos diarios (arroz blanco, arroz coco, frijoles)

    La búsqueda de contornos en menu_items usa normalización de caracteres
    (sin tildes, sin distinción mayúsculas) porque SQLite LIKE no maneja
    correctamente caracteres Unicode > ASCII.
    """
    from app.models.menu_ciclo import MenuCiclo, MenuCicloConfig

    DIAS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    tomorrow = date.today() + timedelta(days=1)
    dia_semana = tomorrow.isoweekday()  # 1=Lun … 7=Dom

    cfg = db.query(MenuCicloConfig).first()
    ciclo_activo = cfg.ciclo_activo if cfg else 1

    row = db.query(MenuCiclo).filter(
        MenuCiclo.ciclo == ciclo_activo,
        MenuCiclo.dia_semana == dia_semana,
    ).first()

    # Pre-cargar todos los menu_items activos con índice normalizado
    todos_items: List[MenuItem] = db.query(MenuItem).filter(MenuItem.activo == True).all()
    idx_items: dict[str, MenuItem] = {_normalizar(it.nombre): it for it in todos_items}

    def buscar_item(nombre: str):
        return idx_items.get(_normalizar(nombre))

    platos = []
    ids_vistos: set = set()

    def add_plato(tipo: str, item_id_val, nombre: str):
        if item_id_val and item_id_val not in ids_vistos:
            platos.append({"tipo": tipo, "item_id": item_id_val, "nombre": nombre})
            ids_vistos.add(item_id_val)

    if row:
        # 1. Sopa principal del ciclo
        if row.sopa_id and row.sopa:
            add_plato("sopa", row.sopa_id, row.sopa.nombre)

        # 1b. Segunda sopa: Consomé de Pescado (Viernes, Sábado, Domingo — confirmado en imágenes)
        if dia_semana in DIAS_CON_CONSOMÉ:
            consomé = buscar_item("CONSOMÉ DE PESCADO")
            if consomé:
                add_plato("sopa", consomé.id, consomé.nombre)

        # 2. Proteína del ciclo
        if row.proteina_id and row.proteina:
            add_plato("proteina", row.proteina_id, row.proteina.nombre)

        # 3. Contornos — campo de texto libre separado por comas.
        #    Normalizamos para comparación robusta con Unicode.
        if row.contorno:
            partes = [p.strip() for p in row.contorno.split(",") if p.strip()]
            for nombre_contorno in partes:
                item = buscar_item(nombre_contorno)
                if item:
                    add_plato("contorno", item.id, item.nombre)
                else:
                    # No encontrado → lo mostramos en UI pero sin item_id (sin receta)
                    platos.append({
                        "tipo": "contorno",
                        "item_id": None,
                        "nombre": nombre_contorno,
                    })

        # 4. Ensalada explícita (ensalada_id), si no fue ya incluida como contorno
        if row.ensalada_id and row.ensalada:
            add_plato("ensalada", row.ensalada_id, row.ensalada.nombre)

    # 5. Platos fijos diarios (siempre, independiente del ciclo)
    for nombre_fijo, subtipo in PLATOS_FIJOS:
        item = buscar_item(nombre_fijo)
        if item and item.id not in ids_vistos:
            platos.append({"tipo": "fijo", "subtipo": subtipo,
                           "item_id": item.id, "nombre": item.nombre})
            ids_vistos.add(item.id)

    return {
        "tomorrow":          str(tomorrow),
        "dia_nombre":        DIAS[dia_semana - 1],
        "ciclo":             ciclo_activo,
        "ciclo_configurado": row is not None,
        "contorno":          row.contorno if row else None,
        "platos":            platos,
    }


# ── Registro masivo de entradas (tras una compra) ─────────────────────────────

class BulkMovementEntry(BaseModel):
    ingredient_id: int
    cantidad: float
    motivo: str = "Compra registrada"


@router.post("/bulk-movements")
def bulk_movements(
    entries: List[BulkMovementEntry],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Registra múltiples entradas de stock de una vez.
    Útil para ingresar una compra completa tras generar la lista de compras.
    """
    count = 0
    for e in entries:
        if e.cantidad <= 0:
            continue
        ing = db.query(Ingredient).filter(Ingredient.id == e.ingredient_id).first()
        if not ing:
            continue
        ing.stock_actual += e.cantidad
        mov = StockMovement(
            ingredient_id=e.ingredient_id,
            tipo="entrada",
            cantidad=e.cantidad,
            motivo=e.motivo,
            usuario_id=current_user.id,
        )
        db.add(mov)
        count += 1
    db.commit()
    return {"ok": True, "registrados": count}
